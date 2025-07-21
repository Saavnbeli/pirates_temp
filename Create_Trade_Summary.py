import os, pprint  
from notion_client import Client
from notion_client.errors import APIResponseError
import time
import pandas as pd
import matplotlib.pyplot as plt
from matplotlib.ticker import FuncFormatter
from openpyxl.styles import Font, PatternFill, Alignment
import plotly.graph_objs as go
from plotly.offline import plot
import requests


# https://www.notion.so/Baez-for-Bednar-233cafa69c5f802e8696e06f99a060d1?source=copy_link
package_id = "233cafa69c5f802e8696e06f99a060d1"
# package_id = None
package_name =  "Baez for Bednar"


notion = Client(auth="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")

# Set Database IDs
DATABASE_ID = "20ecafa69c5f815aadf3fc4cc820f8f6"
DATABASE_ID_PLAYERS = "20ecafa69c5f815aadf3fc4cc820f8f6"
DATABASE_ID_POS = "20ecafa69c5f81aa84f9fc6f95870346"
DATABASE_ID_ORG = "20ecafa69c5f81c2ad7bf3c9ab6871c5"
DATABASE_ID_VALUATION = "212cafa69c5f80a5bec3dce00469cb4e"
DATABASE_ID_PACKAGE = "20ecafa69c5f81a2a006da985e701e07"

# Get schema after database IDs are defined
schema = notion.databases.retrieve(database_id=DATABASE_ID_PLAYERS)["properties"]

def pull_trade_package_data(package_id):
    """
    Pulls data from a specific trade package and all associated players and valuations.
    
    Args:
        package_id (str): The Notion page ID of the trade package
        
    Returns:
        dict: Dictionary containing package data, players, and valuations
    """
    try:
        # Get the trade package data
        package_response = notion.pages.retrieve(page_id=package_id)
        package_data = package_response["properties"]
        
        # Extract package information
        package_info = {
            "package_id": package_id,
            "package_url": package_response["url"],
            "created_time": package_response["created_time"],
            "last_edited_time": package_response["last_edited_time"]
        }
        
        # Extract package properties (adjust based on your actual schema)
        if "Name" in package_data and len(package_data["Name"]["title"]) > 0:
            package_info["name"] = package_data["Name"]["title"][0]["plain_text"]
        
        if "Description" in package_data and len(package_data["Description"]["rich_text"]) > 0:
            package_info["description"] = package_data["Description"]["rich_text"][0]["plain_text"]
        
        if "Notes" in package_data and len(package_data["Notes"]["rich_text"]) > 0:
            package_info["notes"] = package_data["Notes"]["rich_text"][0]["plain_text"]
        
        if "Status" in package_data and package_data["Status"]["select"]:
            package_info["status"] = package_data["Status"]["select"]["name"]
        
        if "Value" in package_data and package_data["Value"]["number"]:
            package_info["value"] = package_data["Value"]["number"]
        
        if "Gain" in package_data and package_data["Gain"]["number"]:
            package_info["gain"] = package_data["Gain"]["number"]
        
        # Get players associated with this package
        players_data = []
        if "Players" in package_data and package_data["Players"]["relation"]:
            player_relations = package_data["Players"]["relation"]
            
            for player_relation in player_relations:
                player_id = player_relation["id"]
                
                # Get player details
                player_response = notion.pages.retrieve(page_id=player_id)
                player_props = player_response["properties"]
                
                player_info = {
                    "player_id": player_id,
                    "player_url": player_response["url"],
                    "created_time": player_response["created_time"],
                    "last_edited_time": player_response["last_edited_time"]
                }
                
                # Extract player properties
                if "Name" in player_props and len(player_props["Name"]["title"]) > 0:
                    # Concatenate all text blocks in the title to handle multi-part names and hyperlinks
                    name_parts = []
                    for text_block in player_props["Name"]["title"]:
                        text = text_block["plain_text"]
                        href = text_block.get("href")
                        if href:
                            name_parts.append(f'<a href="{href}" target="_blank">{text}</a>')
                        else:
                            name_parts.append(text)
                    player_info["name_html"] = "".join(name_parts).strip()
                    player_info["name"] = "".join([tb["plain_text"] for tb in player_props["Name"]["title"]]).strip()
                
                if "PiratesID" in player_props and len(player_props["PiratesID"]["rich_text"]) > 0:
                    player_info["pirates_id"] = player_props["PiratesID"]["rich_text"][0]["plain_text"]
                
                if "PWF Cat" in player_props and player_props["PWF Cat"]["select"]:
                    player_info["pwf_category"] = player_props["PWF Cat"]["select"]["name"]
                
                if "Acquirable" in player_props:
                    player_info["acquirable"] = player_props["Acquirable"]["checkbox"]
                
                if "Age" in player_props and player_props["Age"]["number"]:
                    player_info["age"] = player_props["Age"]["number"]
                
                if "Svc" in player_props and player_props["Svc"]["number"]:
                    player_info["service_time"] = player_props["Svc"]["number"]
                
                if "AVUpdate" in player_props and player_props["AVUpdate"]["rollup"]["number"]:
                    player_info["asset_value"] = player_props["AVUpdate"]["rollup"]["number"]
                
                if "Options" in player_props and len(player_props["Options"]["rich_text"]) > 0:
                    player_info["options"] = player_props["Options"]["rich_text"][0]["plain_text"]
                
                # Get organization info if available
                if "Org" in player_props and player_props["Org"]["relation"]:
                    org_id = player_props["Org"]["relation"][0]["id"]
                    org_response = notion.pages.retrieve(page_id=org_id)
                    if "Name" in org_response["properties"] and len(org_response["properties"]["Name"]["title"]) > 0:
                        player_info["organization"] = org_response["properties"]["Name"]["title"][0]["plain_text"]
                
                # Get position info if available
                if "Pos" in player_props and player_props["Pos"]["relation"]:
                    pos_id = player_props["Pos"]["relation"][0]["id"]
                    pos_response = notion.pages.retrieve(page_id=pos_id)
                    if "Name" in pos_response["properties"] and len(pos_response["properties"]["Name"]["title"]) > 0:
                        player_info["position"] = pos_response["properties"]["Name"]["title"][0]["plain_text"]
                
                # Get valuations for this player
                valuations_data = []
                valuations_response = notion.databases.query(
                    **{
                        "database_id": DATABASE_ID_VALUATION,
                        "filter": {
                            "property": "Players",
                            "relation": {
                                "contains": player_id
                            }
                        }
                    }
                )
                
                for valuation in valuations_response["results"]:
                    #print('here')
                    val_props = valuation["properties"]
                    
                    valuation_info = {
                        "valuation_id": valuation["id"],
                        "valuation_url": valuation["url"],
                        "created_time": valuation["created_time"],
                        "last_edited_time": valuation["last_edited_time"]
                    }
                    
                    # Extract valuation properties
                    if "Name" in val_props and len(val_props["Name"]["title"]) > 0:
                        valuation_info["name"] = val_props["Name"]["title"][0]["plain_text"]
                    
                    if "Year" in val_props and val_props["Year"]["number"]:
                        valuation_info["year"] = val_props["Year"]["number"]
                    
                    if "WAR" in val_props and val_props["WAR"]["number"]:
                        valuation_info["war"] = val_props["WAR"]["number"]
                    
                    if "WAR150" in val_props and val_props["WAR150"]["number"]:
                        valuation_info["war150"] = val_props["WAR150"]["number"]
                    
                    if "WARRoster" in val_props and val_props["WARRoster"]['formula']['number']:
                        valuation_info["war_roster"] = val_props["WARRoster"]['formula']['number']
                    
                    if "MLPA" in val_props and val_props["MLPA"]["number"]:
                        valuation_info["mlpa"] = val_props["MLPA"]["number"]
                    
                    if "Salary Raw" in val_props and val_props["Salary Raw"]["number"]:
                        valuation_info["salary"] = val_props["Salary Raw"]["number"]
                    
                    if "AV" in val_props and val_props["AV"]["formula"]["number"]:
                        valuation_info["asset_value"] = val_props["AV"]["formula"]["number"]
                    
                    if "Adjustment" in val_props and val_props["Adjustment"]["number"]:
                        valuation_info["adjustment"] = val_props["Adjustment"]["number"]
                    
                    if "RosterFit" in val_props and val_props["RosterFit"]["number"]:
                        valuation_info["roster_fit"] = val_props["RosterFit"]["number"]
                    
                    if "LastGuaranteedYr" in val_props and val_props["LastGuaranteedYr"]["number"]:
                        valuation_info["last_guaranteed_year"] = val_props["LastGuaranteedYr"]["number"]
                    
                    valuations_data.append(valuation_info)
                
                player_info["valuations"] = valuations_data
                players_data.append(player_info)
        
        # Compile all data
        trade_package_data = {
            "package": package_info,
            "players": players_data,
            "total_players": len(players_data),
            "total_valuations": sum(len(player["valuations"]) for player in players_data)
        }
        
        return trade_package_data
        
    except APIResponseError as e:
        print(f"API Error: {e}")
        return None
    except Exception as e:
        print(f"Error pulling trade package data: {e}")
        return None

def get_all_trade_packages():
    """
    Gets all trade packages from the DATABASE_ID_PACKAGE database.
    
    Returns:
        list: List of trade package summaries
    """
    try:
        packages = []
        start_cursor = None
        
        while True:
            response = notion.databases.query(
                **{
                    "database_id": DATABASE_ID_PACKAGE,
                    "start_cursor": start_cursor
                }
            )
            
            for page in response["results"]:
                package_summary = {
                    "package_id": page["id"],
                    "package_url": page["url"],
                    "created_time": page["created_time"],
                    "last_edited_time": page["last_edited_time"]
                }
                
                # Extract basic properties
                props = page["properties"]
                if "Name" in props and len(props["Name"]["title"]) > 0:
                    package_summary["name"] = props["Name"]["title"][0]["plain_text"]
                
                if "Description" in props and len(props["Description"]["rich_text"]) > 0:
                    package_summary["description"] = props["Description"]["rich_text"][0]["plain_text"]
                
                if "Status" in props and props["Status"]["select"]:
                    package_summary["status"] = props["Status"]["select"]["name"]
                
                if "Value" in props and props["Value"]["number"]:
                    package_summary["value"] = props["Value"]["number"]
                
                if "Gain" in props and props["Gain"]["number"]:
                    package_summary["gain"] = props["Gain"]["number"]
                
                # Count players in package
                if "Players" in props and props["Players"]["relation"]:
                    package_summary["player_count"] = len(props["Players"]["relation"])
                else:
                    package_summary["player_count"] = 0
                
                packages.append(package_summary)
            
            start_cursor = response.get("next_cursor")
            if not start_cursor:
                break
        
        return packages
        
    except APIResponseError as e:
        print(f"API Error: {e}")
        return []
    except Exception as e:
        print(f"Error getting trade packages: {e}")
        return []

def analyze_trade_package(package_data):
    """
    Analyzes a trade package and provides summary statistics.
    
    Args:
        package_data (dict): Output from pull_trade_package_data function
        
    Returns:
        dict: Analysis summary
    """
    if not package_data:
        return None
    
    analysis = {
        "package_name": package_data["package"].get("name", "Unknown"),
        "total_players": package_data["total_players"],
        "total_valuations": package_data["total_valuations"],
        "player_summary": {},
        "valuation_summary": {}
    }
    
    # Player analysis
    if package_data["players"]:
        positions = {}
        organizations = {}
        age_range = {"min": float('inf'), "max": float('-inf')}
        total_asset_value = 0
        
        for player in package_data["players"]:
            # Position breakdown
            pos = player.get("position", "Unknown")
            positions[pos] = positions.get(pos, 0) + 1
            
            # Organization breakdown
            org = player.get("organization", "Unknown")
            organizations[org] = organizations.get(org, 0) + 1
            
            # Age range
            if player.get("age"):
                age = player["age"]
                age_range["min"] = min(age_range["min"], age)
                age_range["max"] = max(age_range["max"], age)
            
            # Asset value
            if player.get("asset_value"):
                total_asset_value += player["asset_value"]
        
        analysis["player_summary"] = {
            "positions": positions,
            "organizations": organizations,
            "age_range": age_range if age_range["min"] != float('inf') else None,
            "total_asset_value": total_asset_value
        }
    
    # Valuation analysis
    all_valuations = []
    for player in package_data["players"]:
        all_valuations.extend(player.get("valuations", []))
    
    if all_valuations:
        years = set()
        total_war = 0
        total_war_roster = 0
        total_salary = 0
        total_AV = 0
        
        for val in all_valuations:
            if val.get("year"):
                years.add(val["year"])
            if val.get("war"):
                total_war += val["war"]
            if val.get("war_roster"):
                total_war_roster += val["war_roster"]
            if val.get("salary"):
                total_salary += val["salary"]
            if val.get("asset_value"):
                total_AV += val["asset_value"]
        analysis["valuation_summary"] = {
            "years_covered": sorted(list(years)),
            "total_war": total_war,
            "total_war_roster": total_war_roster,
            "total_salary": total_salary,
            "total_AV": total_AV,
            "avg_war_per_player": total_war / len(package_data["players"]) if package_data["players"] else 0,
            "avg_war_roster_per_player": total_war_roster / len(package_data["players"]) if package_data["players"] else 0
        }
    
    return analysis

def print_trade_package_summary(package_data, analysis=None):
    """
    Prints a formatted summary of a trade package.
    
    Args:
        package_data (dict): Output from pull_trade_package_data function
        analysis (dict): Optional output from analyze_trade_package function
    """
    if not package_data:
        print("No package data available")
        return
    
    print("=" * 60)
    print(f"TRADE PACKAGE: {package_data['package'].get('name', 'Unknown')}")
    print("=" * 60)
    print(f"Package ID: {package_data['package']['package_id']}")
    print(f"URL: {package_data['package']['package_url']}")
    print(f"Players: {package_data['total_players']}")
    print(f"Valuations: {package_data['total_valuations']}")
    
    if package_data['package'].get('description'):
        print(f"Description: {package_data['package']['description']}")
    
    if package_data['package'].get('status'):
        print(f"Status: {package_data['package']['status']}")
    
    if package_data['package'].get('value'):
        print(f"Value: {package_data['package']['value']}")
    
    if package_data['package'].get('gain'):
        print(f"Gain: {package_data['package']['gain']}")
    
    print("\n" + "-" * 40)
    print("PLAYERS:")
    print("-" * 40)
    
    for i, player in enumerate(package_data["players"], 1):
        print(f"{i}. {player.get('name', 'Unknown')}")
        print(f"   Position: {player.get('position', 'Unknown')}")
        print(f"   Organization: {player.get('organization', 'Unknown')}")
        print(f"   Age: {player.get('age', 'Unknown')}")
        print(f"   Asset Value: {player.get('asset_value', 'Unknown')}")
        print(f"   Valuations: {len(player.get('valuations', []))}")
        print()
    
    if analysis:
        print("-" * 40)
        print("ANALYSIS:")
        print("-" * 40)
        
        if analysis["player_summary"]:
            ps = analysis["player_summary"]
            print(f"Position Breakdown: {ps['positions']}")
            print(f"Organization Breakdown: {ps['organizations']}")
            if ps['age_range']:
                print(f"Age Range: {ps['age_range']['min']} - {ps['age_range']['max']}")
            print(f"Total Asset Value: {ps['total_asset_value']}")
        
        if analysis["valuation_summary"]:
            vs = analysis["valuation_summary"]
            print(f"Years Covered: {vs['years_covered']}")
            print(f"Total WAR (Projected): {vs['total_war']:.2f}")
            print(f"Total WAR (Roster): {vs['total_war_roster']:.2f}")
            print(f"Total Salary: {vs['total_salary']:,.0f}")
            print(f"Total Asset Value: {vs['total_AV']:,.0f}")
            print(f"Average WAR per Player: {vs['avg_war_per_player']:.2f}")
            print(f"Average WAR Roster per Player: {vs['avg_war_roster_per_player']:.2f}")
    
    print("=" * 60)

# Example usage functions
def example_get_package_by_name(package_name):
    """
    Example: Get a trade package by searching for its name.
    
    Args:
        package_name (str): Name of the trade package to find
        
    Returns:
        dict: Trade package data or None if not found
    """
    all_packages = get_all_trade_packages()
    
    for package in all_packages:
        if package.get("name") == package_name:
            return pull_trade_package_data(package["package_id"])
    
    print(f"Package '{package_name}' not found")
    return None

def example_get_packages_by_status(status):
    """
    Example: Get all trade packages with a specific status.
    
    Args:
        status (str): Status to filter by
        
    Returns:
        list: List of trade package data
    """
    all_packages = get_all_trade_packages()
    matching_packages = []
    
    for package in all_packages:
        if package.get("status") == status:
            full_data = pull_trade_package_data(package["package_id"])
            if full_data:
                matching_packages.append(full_data)
    
    return matching_packages

# pkg = pull_trade_package_data("21fcafa69c5f808eb804f63d55568c56")
# pkg_analysis = analyze_trade_package(pkg)
# print_trade_package_summary(pkg, pkg_analysis)

def compare_war_projections_vs_roster(package_data):
    """
    Compares WAR projections vs roster WAR for a trade package and provides detailed analysis.
    
    Args:
        package_data (dict): Output from pull_trade_package_data function
        
    Returns:
        dict: Comparison analysis
    """
    if not package_data or not package_data["players"]:
        return None
    
    comparison = {
        "package_name": package_data["package"].get("name", "Unknown"),
        "total_players": len(package_data["players"]),
        "war_comparison": {},
        "player_details": []
    }
    
    # Aggregate WAR data by year
    year_data = {}
    
    for player in package_data["players"]:
        player_comparison = {
            "player_name": player.get("name", "Unknown"),
            "position": player.get("position", "Unknown"),
            "organization": player.get("organization", "Unknown"),
            "valuations": []
        }
        
        for val in player.get("valuations", []):
            year = val.get("year")
            if year:
                if year not in year_data:
                    year_data[year] = {
                        "war_proj": 0,
                        "war_roster": 0,
                        "asset_value": 0,
                        "salary": 0,
                        "player_count": 0
                    }
                
                war_proj = val.get("war", 0)
                war_roster = val.get("war_roster", 0)
                asset_value = val.get("asset_value", 0)
                salary = val.get("salary", 0)
                
                year_data[year]["war_proj"] += war_proj
                year_data[year]["war_roster"] += war_roster
                year_data[year]["asset_value"] += asset_value
                year_data[year]["salary"] += salary
                year_data[year]["player_count"] += 1
                
                player_comparison["valuations"].append({
                    "year": year,
                    "war_proj": war_proj,
                    "war_roster": war_roster,
                    "war_difference": war_proj - war_roster,
                    "asset_value": asset_value,
                    "salary": salary
                })
        
        comparison["player_details"].append(player_comparison)
    
    # Calculate overall statistics
    total_war_proj = sum(data["war_proj"] for data in year_data.values())
    total_war_roster = sum(data["war_roster"] for data in year_data.values())
    total_asset_value = sum(data["asset_value"] for data in year_data.values())
    total_salary = sum(data["salary"] for data in year_data.values())
    
    comparison["war_comparison"] = {
        "by_year": year_data,
        "total_war_proj": total_war_proj,
        "total_war_roster": total_war_roster,
        "total_war_difference": total_war_proj - total_war_roster,
        "war_difference_percentage": ((total_war_proj - total_war_roster) / total_war_roster * 100) if total_war_roster != 0 else 0,
        "total_asset_value": total_asset_value,
        "total_salary": total_salary,
        "avg_war_proj_per_player": total_war_proj / len(package_data["players"]) if package_data["players"] else 0,
        "avg_war_roster_per_player": total_war_roster / len(package_data["players"]) if package_data["players"] else 0
    }
    
    return comparison

def print_war_comparison_summary(comparison):
    """
    Prints a detailed summary of WAR projections vs roster WAR comparison.
    
    Args:
        comparison (dict): Output from compare_war_projections_vs_roster function
    """
    if not comparison:
        print("No comparison data available")
        return
    
    print("=" * 80)
    print(f"WAR PROJECTIONS vs ROSTER WAR COMPARISON")
    print(f"Package: {comparison['package_name']}")
    print("=" * 80)
    
    war_comp = comparison["war_comparison"]
    
    print(f"Total Players: {comparison['total_players']}")
    print(f"Total WAR (Projected): {war_comp['total_war_proj']:.2f}")
    print(f"Total WAR (Roster): {war_comp['total_war_roster']:.2f}")
    print(f"Difference: {war_comp['total_war_difference']:.2f}")
    print(f"Difference %: {war_comp['war_difference_percentage']:.1f}%")
    print(f"Total Asset Value: {war_comp['total_asset_value']:,.0f}")
    print(f"Total Salary: {war_comp['total_salary']:,.0f}")
    print()
    
    print(f"Average WAR per Player (Projected): {war_comp['avg_war_proj_per_player']:.2f}")
    print(f"Average WAR per Player (Roster): {war_comp['avg_war_roster_per_player']:.2f}")
    print()
    
    # Analyze PIT vs non-PIT players
    pit_players = []
    non_pit_players = []
    
    for player in comparison["player_details"]:
        if player.get("organization") == "PIT":
            pit_players.append(player)
        else:
            non_pit_players.append(player)
    
    print("-" * 60)
    print("PIT vs NON-PIT PLAYER ANALYSIS:")
    print("-" * 60)
    print(f"PIT Players: {len(pit_players)}")
    print(f"Non-PIT Players: {len(non_pit_players)}")
    print()
    
    # Calculate totals by organization
    pit_war_roster = 0
    non_pit_war_roster = 0
    pit_war_proj = 0
    non_pit_war_proj = 0
    pit_salary = 0
    non_pit_salary = 0
    pit_av = 0
    non_pit_av = 0
    
    for player in pit_players:
        for val in player["valuations"]:
            pit_war_roster += val.get("war_roster", 0)
            pit_war_proj += val.get("war_proj", 0)
            pit_salary += val.get("salary", 0)
            pit_av += val.get("asset_value", 0)
    
    for player in non_pit_players:
        for val in player["valuations"]:
            non_pit_war_roster += val.get("war_roster", 0)
            non_pit_war_proj += val.get("war_proj", 0)
            non_pit_salary += val.get("salary", 0)
            non_pit_av += val.get("asset_value", 0)
    
    print(f"PIT Players - Total WAR (Roster): {pit_war_roster:.2f}")
    print(f"PIT Players - Total WAR (Projected): {pit_war_proj:.2f}")
    print(f"PIT Players - Total Salary: ${pit_salary:,.0f}")
    print(f"PIT Players - Total AV: ${pit_av:,.0f}")
    print(f"Non-PIT Players - Total WAR (Roster): {non_pit_war_roster:.2f}")
    print(f"Non-PIT Players - Total WAR (Projected): {non_pit_war_proj:.2f}")
    print(f"Non-PIT Players - Total Salary: ${non_pit_salary:,.0f}")
    print(f"Non-PIT Players - Total AV: ${non_pit_av:,.0f}")
    print()
    
    # Net changes
    net_war_roster = pit_war_roster - non_pit_war_roster
    net_war_proj = pit_war_proj - non_pit_war_proj
    net_salary = pit_salary - non_pit_salary
    net_av = pit_av - non_pit_av
    
    print(f"NET CHANGES (PIT - Non-PIT):")
    print(f"  WAR (Roster): {net_war_roster:+.2f}")
    print(f"  WAR (Projected): {net_war_proj:+.2f}")
    print(f"  Salary: ${net_salary:+,.0f}")
    print(f"  AV: ${net_av:+,.0f}")
    print()
    
    # Year-by-year breakdown with PIT vs non-PIT
    print("-" * 60)
    print("YEAR-BY-YEAR BREAKDOWN (PIT vs NON-PIT):")
    print("-" * 60)
    
    # Group valuations by year and organization
    year_org_data = {}
    
    for player in comparison["player_details"]:
        org = player.get("organization", "Unknown")
        for val in player["valuations"]:
            year = val.get("year")
            if year:
                if year not in year_org_data:
                    year_org_data[year] = {
                        "PIT": {"war_roster": 0, "war_proj": 0, "salary": 0, "av": 0, "players": 0}, 
                        "non_PIT": {"war_roster": 0, "war_proj": 0, "salary": 0, "av": 0, "players": 0}
                    }
                
                if org == "PIT":
                    year_org_data[year]["PIT"]["war_roster"] += val.get("war_roster", 0)
                    year_org_data[year]["PIT"]["war_proj"] += val.get("war_proj", 0)
                    year_org_data[year]["PIT"]["salary"] += val.get("salary", 0)
                    year_org_data[year]["PIT"]["av"] += val.get("asset_value", 0)
                    year_org_data[year]["PIT"]["players"] += 1
                else:
                    year_org_data[year]["non_PIT"]["war_roster"] += val.get("war_roster", 0)
                    year_org_data[year]["non_PIT"]["war_proj"] += val.get("war_proj", 0)
                    year_org_data[year]["non_PIT"]["salary"] += val.get("salary", 0)
                    year_org_data[year]["non_PIT"]["av"] += val.get("asset_value", 0)
                    year_org_data[year]["non_PIT"]["players"] += 1
    
    for year in sorted(year_org_data.keys()):
        pit_data = year_org_data[year]["PIT"]
        non_pit_data = year_org_data[year]["non_PIT"]
        
        print(f"Year {year}:")
        print(f"  PIT Players ({pit_data['players']}):")
        print(f"    WAR Roster: {pit_data['war_roster']:.2f}")
        print(f"    WAR Projected: {pit_data['war_proj']:.2f}")
        print(f"    Roster vs Proj Diff: {pit_data['war_roster'] - pit_data['war_proj']:+.2f}")
        print(f"    Salary: ${pit_data['salary']:,.0f}")
        print(f"    AV: ${pit_data['av']:,.0f}")
        
        print(f"  Non-PIT Players ({non_pit_data['players']}):")
        print(f"    WAR Roster: {non_pit_data['war_roster']:.2f}")
        print(f"    WAR Projected: {non_pit_data['war_proj']:.2f}")
        print(f"    Roster vs Proj Diff: {non_pit_data['war_roster'] - non_pit_data['war_proj']:+.2f}")
        print(f"    Salary: ${non_pit_data['salary']:,.0f}")
        print(f"    AV: ${non_pit_data['av']:,.0f}")
        
        # Calculate net change for this year
        net_roster_change = pit_data['war_roster'] - non_pit_data['war_roster']
        net_proj_change = pit_data['war_proj'] - non_pit_data['war_proj']
        net_salary_change = pit_data['salary'] - non_pit_data['salary']
        net_av_change = pit_data['av'] - non_pit_data['av']
        
        print(f"  Net Change (PIT - Non-PIT):")
        print(f"    WAR Roster: {net_roster_change:+.2f}")
        print(f"    WAR Projected: {net_proj_change:+.2f}")
        print(f"    Salary: ${net_salary_change:+,.0f}")
        print(f"    AV: ${net_av_change:+,.0f}")
        print()
    
    print("-" * 60)
    print("PLAYER DETAILS:")
    print("-" * 60)
    
    # Group players by organization
    print("PIT PLAYERS:")
    for player in pit_players:
        print(f"  {player['player_name']} ({player['position']})")
        for val in player["valuations"]:
            print(f"    {val['year']}: Proj={val['war_proj']:.2f}, Roster={val['war_roster']:.2f}, Diff={val['war_difference']:+.2f}")
        print()
    
    print("NON-PIT PLAYERS:")
    for player in non_pit_players:
        print(f"  {player['player_name']} ({player['position']}) - {player['organization']}")
        for val in player["valuations"]:
            print(f"    {val['year']}: Proj={val['war_proj']:.2f}, Roster={val['war_roster']:.2f}, Diff={val['war_difference']:+.2f}")
        print()
    
    print("=" * 80)



def plot_WAR_Change_Over_Time(package_data):
    """
    Plots the WAR change over time for a trade package.
    
    Args:
        package_data (dict): Output from pull_trade_package_data function
    """ 
    # Create a figure and axis
    fig, ax = plt.subplots(figsize=(12, 8))
    
    # Get all valuations
    all_valuations = []
    for player in package_data["players"]:
        all_valuations.extend(player.get("valuations", [])) 

    # Extract years and WAR values
    years = []
    war_proj_values = []
    war_roster_values = []
    asset_values = []
    
    for val in all_valuations:
        if val.get("year"):
            years.append(val["year"])
            war_proj_values.append(val.get("war", 0))
            war_roster_values.append(val.get("war_roster", 0))
            asset_values.append(val.get("asset_value", 0))
    
    # Create subplots
    fig, (ax1, ax2) = plt.subplots(2, 1, figsize=(12, 10))
    
    # Plot WAR projections vs Roster WAR
    ax1.plot(years, war_proj_values, 'b-o', label='WAR Projected', linewidth=2, markersize=6)
    ax1.plot(years, war_roster_values, 'r-s', label='WAR Roster', linewidth=2, markersize=6)
    ax1.set_xlabel('Year')
    ax1.set_ylabel('WAR')
    ax1.set_title(f'WAR Projections vs Roster WAR - {package_data["package"].get("name", "Trade Package")}')
    ax1.legend()
    ax1.grid(True, alpha=0.3)
    
    # Plot Asset Values
    ax2.plot(years, asset_values, 'g-^', label='Asset Value', linewidth=2, markersize=6)
    ax2.set_xlabel('Year')
    ax2.set_ylabel('Asset Value')
    ax2.set_title(f'Asset Value Over Time - {package_data["package"].get("name", "Trade Package")}')
    ax2.legend()
    ax2.grid(True, alpha=0.3)
    
    plt.tight_layout()
    plt.show()
    
    return fig

def plot_metric_changes_over_time(package_data, metric='war_roster', figsize=(12, 8)):
    """
    Plots specified metric changes over time for a trade package.
    
    Args:
        package_data (dict): Output from pull_trade_package_data function
        metric (str): Metric to plot - 'war_roster', 'war_proj', 'salary', or 'av'
        figsize (tuple): Figure size (width, height)
        
    Returns:
        matplotlib.figure.Figure: The created figure
    """
    if not package_data or not package_data["players"]:
        print("No package data available for plotting")
        return None
    
    # Validate metric parameter
    valid_metrics = ['war_roster', 'war_proj', 'salary', 'av']
    if metric not in valid_metrics:
        print(f"Invalid metric '{metric}'. Valid options: {valid_metrics}")
        return None
    
    # Create figure and axis
    fig, ax = plt.subplots(figsize=figsize)
    
    # Get all valuations
    all_valuations = []
    for player in package_data["players"]:
        all_valuations.extend(player.get("valuations", []))
    
    if not all_valuations:
        print("No valuation data available for plotting")
        return None
    
    # Group data by year and organization
    year_org_data = {}
    
    for player in package_data["players"]:
        org = player.get("organization", "Unknown")
        for val in player.get("valuations", []):
            year = val.get("year")
            if year:
                if year not in year_org_data:
                    year_org_data[year] = {"PIT": 0, "non_PIT": 0}
                
                # Get the appropriate metric value
                if metric == 'war_roster':
                    value = val.get("war_roster", 0)
                elif metric == 'war_proj':
                    value = val.get("war", 0)
                elif metric == 'salary':
                    value = val.get("salary", 0)
                elif metric == 'av':
                    value = val.get("asset_value", 0)
                
                if org == "PIT":
                    year_org_data[year]["PIT"] += value
                else:
                    year_org_data[year]["non_PIT"] += value
    
    # Prepare data for plotting
    years = sorted(year_org_data.keys())
    pit_values = [year_org_data[year]["PIT"] for year in years]
    non_pit_values = [year_org_data[year]["non_PIT"] for year in years]
    net_values = [pit_values[i] - non_pit_values[i] for i in range(len(years))]
    
    # Create the plot
    ax.plot(years, pit_values, 'b-o', label='PIT Players', linewidth=2, markersize=8)
    ax.plot(years, non_pit_values, 'r-s', label='Non-PIT Players', linewidth=2, markersize=8)
    ax.plot(years, net_values, 'g-^', label='Net Change (PIT - Non-PIT)', linewidth=3, markersize=10)
    
    # Customize the plot
    metric_labels = {
        'war_roster': 'WAR Roster',
        'war_proj': 'WAR Projected',
        'salary': 'Salary ($)',
        'av': 'Asset Value ($)'
    }
    
    ax.set_xlabel('Year', fontsize=12)
    ax.set_ylabel(metric_labels[metric])
    ax.set_title(f'{metric_labels[metric]} Changes Over Time - {package_data["package"].get("name", "Trade Package")}', 
                 fontsize=14, fontweight='bold')
    ax.legend(fontsize=11)
    ax.grid(True, alpha=0.3)
    
    # Format y-axis for currency if needed
    if metric in ['salary', 'av']:
        ax.yaxis.set_major_formatter(FuncFormatter(lambda x, p: f'${x:,.0f}'))
    
    # Add value annotations
    for i, year in enumerate(years):
        # Annotate PIT values
        ax.annotate(f'{pit_values[i]:.1f}', 
                   (year, pit_values[i]), 
                   textcoords="offset points", 
                   xytext=(0,10), 
                   ha='center', 
                   fontsize=9)
        
        # Annotate non-PIT values
        ax.annotate(f'{non_pit_values[i]:.1f}', 
                   (year, non_pit_values[i]), 
                   textcoords="offset points", 
                   xytext=(0,-15), 
                   ha='center', 
                   fontsize=9)
        
        # Annotate net values
        ax.annotate(f'{net_values[i]:.1f}', 
                   (year, net_values[i]), 
                   textcoords="offset points", 
                   xytext=(0,10), 
                   ha='center', 
                   fontsize=9,
                   fontweight='bold')
    
    plt.tight_layout()
    plt.show()
    
    return fig

def plot_all_metrics_comparison(package_data, figsize=(15, 12)):
    """
    Creates a comprehensive comparison plot showing all metrics over time.
    
    Args:
        package_data (dict): Output from pull_trade_package_data function
        figsize (tuple): Figure size (width, height)
        
    Returns:
        matplotlib.figure.Figure: The created figure
    """
    if not package_data or not package_data["players"]:
        print("No package data available for plotting")
        return None
    
    # Create subplots for all metrics
    fig, axes = plt.subplots(2, 2, figsize=figsize)
    fig.suptitle(f'Comprehensive Trade Package Analysis - {package_data["package"].get("name", "Trade Package")}', 
                 fontsize=16, fontweight='bold')
    
    metrics = ['war_roster', 'war_proj', 'salary', 'av']
    metric_labels = {
        'war_roster': 'WAR Roster',
        'war_proj': 'WAR Projected', 
        'salary': 'Salary ($)',
        'av': 'Asset Value ($)'
    }
    
    # Get all valuations
    all_valuations = []
    for player in package_data["players"]:
        all_valuations.extend(player.get("valuations", []))
    
    if not all_valuations:
        print("No valuation data available for plotting")
        return None
    
    # Group data by year and organization
    year_org_data = {}
    
    for player in package_data["players"]:
        org = player.get("organization", "Unknown")
        for val in player.get("valuations", []):
            year = val.get("year")
            if year:
                if year not in year_org_data:
                    year_org_data[year] = {
                        "PIT": {"war_roster": 0, "war_proj": 0, "salary": 0, "av": 0},
                        "non_PIT": {"war_roster": 0, "war_proj": 0, "salary": 0, "av": 0}
                    }
                
                if org == "PIT":
                    year_org_data[year]["PIT"]["war_roster"] += val.get("war_roster", 0)
                    year_org_data[year]["PIT"]["war_proj"] += val.get("war", 0)
                    year_org_data[year]["PIT"]["salary"] += val.get("salary", 0)
                    year_org_data[year]["PIT"]["av"] += val.get("asset_value", 0)
                else:
                    year_org_data[year]["non_PIT"]["war_roster"] += val.get("war_roster", 0)
                    year_org_data[year]["non_PIT"]["war_proj"] += val.get("war", 0)
                    year_org_data[year]["non_PIT"]["salary"] += val.get("salary", 0)
                    year_org_data[year]["non_PIT"]["av"] += val.get("asset_value", 0)
    
    years = sorted(year_org_data.keys())
    
    # Plot each metric
    for i, metric in enumerate(metrics):
        ax = axes[i//2, i%2]
        
        pit_values = [year_org_data[year]["PIT"][metric] for year in years]
        non_pit_values = [year_org_data[year]["non_PIT"][metric] for year in years]
        net_values = [pit_values[j] - non_pit_values[j] for j in range(len(years))]
        
        ax.plot(years, pit_values, 'b-o', label='PIT Players', linewidth=2, markersize=6)
        ax.plot(years, non_pit_values, 'r-s', label='Non-PIT Players', linewidth=2, markersize=6)
        ax.plot(years, net_values, 'g-^', label='Net Change', linewidth=2, markersize=6)
        
        ax.set_xlabel('Year')
        ax.set_ylabel(metric_labels[metric])
        ax.set_title(metric_labels[metric])
        ax.legend()
        ax.grid(True, alpha=0.3)
        
        # Format y-axis for currency if needed
        if metric in ['salary', 'av']:
            ax.yaxis.set_major_formatter(FuncFormatter(lambda x, p: f'${x:,.0f}'))
    
    plt.tight_layout()
    plt.show()
    
    return fig

def create_summary_table(package_data, save_to_excel=False, filename=None, output_html=True):
    """
    Creates a summary table in Excel-like format for a trade package.
    
    Args:
        package_data (dict): Output from pull_trade_package_data function
        save_to_excel (bool): Whether to save the table to an Excel file
        filename (str): Optional filename for Excel export
        output_html (bool): Whether to output as formatted HTML
        
    Returns:
        pandas.DataFrame: Formatted summary table
    """
    if not package_data or not package_data["players"]:
        print("No package data available for table creation")
        return None
    
    # Analyze PIT vs non-PIT players
    pit_players = []
    non_pit_players = []
    
    for player in package_data["players"]:
        if player.get("organization") == "PIT":
            pit_players.append(player)
        else:
            non_pit_players.append(player)
    
    # Group valuations by year and organization
    year_org_data = {}
    
    for player in package_data["players"]:
        org = player.get("organization", "Unknown")
        for val in player.get("valuations", []):
            year = val.get("year")
            if year:
                if year not in year_org_data:
                    year_org_data[year] = {
                        "PIT": {"war_roster": 0, "war_proj": 0, "salary": 0, "av": 0, "players": 0}, 
                        "non_PIT": {"war_roster": 0, "war_proj": 0, "salary": 0, "av": 0, "players": 0}
                    }
                
                if org == "PIT":
                    year_org_data[year]["PIT"]["war_roster"] += val.get("war_roster", 0)
                    year_org_data[year]["PIT"]["war_proj"] += val.get("war", 0)
                    year_org_data[year]["PIT"]["salary"] += val.get("salary", 0)
                    year_org_data[year]["PIT"]["av"] += val.get("asset_value", 0)
                    year_org_data[year]["PIT"]["players"] += 1
                else:
                    year_org_data[year]["non_PIT"]["war_roster"] += val.get("war_roster", 0)
                    year_org_data[year]["non_PIT"]["war_proj"] += val.get("war", 0)
                    year_org_data[year]["non_PIT"]["salary"] += val.get("salary", 0)
                    year_org_data[year]["non_PIT"]["av"] += val.get("asset_value", 0)
                    year_org_data[year]["non_PIT"]["players"] += 1
    
    # Create summary table data
    table_data = []
    
    # Package overview section
    table_data.append({
        'Category': 'PACKAGE OVERVIEW',
        'Metric': 'Package Name',
        'Value': package_data["package"].get("name", "Unknown"),
        'Details': f"Total Players: {len(package_data['players'])}"
    })
    
    table_data.append({
        'Category': 'PACKAGE OVERVIEW',
        'Metric': 'PIT Players',
        'Value': len(pit_players),
        'Details': f"Non-PIT Players: {len(non_pit_players)}"
    })
    
    if package_data["package"].get("gain"):
        table_data.append({
            'Category': 'PACKAGE OVERVIEW',
            'Metric': 'Package Gain',
            'Value': f"${package_data['package']['gain']:,.0f}",
            'Details': ''
        })
    
    table_data.append({
        'Category': 'PACKAGE OVERVIEW',
        'Metric': 'Package URL',
        'Value': package_data["package"]["package_url"],
        'Details': ''
    })
    
    # Add empty row for spacing
    table_data.append({
        'Category': '',
        'Metric': '',
        'Value': '',
        'Details': ''
    })
    
    # Year-by-year breakdown
    for year in sorted(year_org_data.keys()):
        pit_data = year_org_data[year]["PIT"]
        non_pit_data = year_org_data[year]["non_PIT"]
        
        # Net calculations
        net_war_roster = pit_data['war_roster'] - non_pit_data['war_roster']
        net_war_proj = pit_data['war_proj'] - non_pit_data['war_proj']
        net_salary = pit_data['salary'] - non_pit_data['salary']
        net_av = pit_data['av'] - non_pit_data['av']
        
        # Year header
        table_data.append({
            'Category': f'YEAR {year}',
            'Metric': 'PIT Players',
            'Value': f"{pit_data['players']} players",
            'Details': f"Non-PIT: {non_pit_data['players']} players"
        })
        
        # WAR Roster
        table_data.append({
            'Category': f'YEAR {year}',
            'Metric': 'WAR Roster',
            'Value': f"PIT: {pit_data['war_roster']:.2f} | Non-PIT: {non_pit_data['war_roster']:.2f}",
            'Details': f"Net: {net_war_roster:+.2f}"
        })
        
        # WAR Projected
        table_data.append({
            'Category': f'YEAR {year}',
            'Metric': 'WAR Projected',
            'Value': f"PIT: {pit_data['war_proj']:.2f} | Non-PIT: {non_pit_data['war_proj']:.2f}",
            'Details': f"Net: {net_war_proj:+.2f}"
        })
        
        # Salary
        table_data.append({
            'Category': f'YEAR {year}',
            'Metric': 'Salary',
            'Value': f"PIT: ${pit_data['salary']:,.0f} | Non-PIT: ${non_pit_data['salary']:,.0f}",
            'Details': f"Net (PIT - Non-PIT): ${net_salary:+,.0f}"
        })
        
        # Asset Value
        table_data.append({
            'Category': f'YEAR {year}',
            'Metric': 'Asset Value',
            'Value': f"PIT: ${pit_data['av']:,.0f} | Non-PIT: ${non_pit_data['av']:,.0f}",
            'Details': f"Net (PIT - Non-PIT): ${net_av:+,.0f}"
        })
        
        # Add empty row for spacing
        table_data.append({
            'Category': '',
            'Metric': '',
            'Value': '',
            'Details': ''
        })
    
    # Player details section
    table_data.append({
        'Category': 'PLAYER DETAILS',
        'Metric': 'PIT PLAYERS',
        'Value': '',
        'Details': ''
    })
    
    for player in pit_players:
        player_valuations = []
        for val in player.get("valuations", []):
            player_valuations.append(f"{val['year']}: {val.get('war_roster', 0):.1f} WAR")
        
        table_data.append({
            'Category': 'PLAYER DETAILS',
            'Metric': f"{player['name']} ({player.get('position', 'Unknown')})",
            'Value': f"AV: ${player.get('asset_value', 0):,.0f}",
            'Details': ' | '.join(player_valuations) if player_valuations else 'No valuations'
        })
    
    table_data.append({
        'Category': 'PLAYER DETAILS',
        'Metric': 'NON-PIT PLAYERS',
        'Value': '',
        'Details': ''
    })
    
    for player in non_pit_players:
        player_valuations = []
        for val in player.get("valuations", []):
            player_valuations.append(f"{val['year']}: {val.get('war_roster', 0):.1f} WAR")
        
        table_data.append({
            'Category': 'PLAYER DETAILS',
            'Metric': f"{player['name']} ({player.get('position', 'Unknown')}) - {player.get('organization', 'Unknown')}",
            'Value': f"AV: ${player.get('asset_value', 0):,.0f}",
            'Details': ' | '.join(player_valuations) if player_valuations else 'No valuations'
        })
    
    # Create DataFrame
    df = pd.DataFrame(table_data)
    
    # Output HTML table if requested
    if output_html:
        html_output = create_html_table(df, package_data["package"].get("name", "Unknown"))
        print(html_output)
        
        # Save HTML to file
        html_filename = f"trade_package_summary_{package_data['package'].get('name', 'unknown').replace(' ', '_')}.html"
        with open(html_filename, 'w', encoding='utf-8') as f:
            f.write(html_output)
        print(f"\nHTML table saved to: {html_filename}")
    
    # Save to Excel if requested
    if save_to_excel:
        if filename is None:
            filename = f"trade_package_summary_{package_data['package'].get('name', 'unknown').replace(' ', '_')}.xlsx"
        
        # Create Excel writer
        with pd.ExcelWriter(filename, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='Summary', index=False)
            
            # Get the workbook and worksheet
            workbook = writer.book
            worksheet = writer.sheets['Summary']
            
            # Auto-adjust column widths
            for column in worksheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                worksheet.column_dimensions[column_letter].width = adjusted_width
            
            # Simple formatting
            from openpyxl.styles import Font, PatternFill
            
            # Header formatting - just bold
            for cell in worksheet[1]:
                cell.font = Font(bold=True)
            
            # Category rows - just bold
            for row_num, row in enumerate(worksheet.iter_rows(min_row=2), start=2):
                if row[0].value in ['PACKAGE OVERVIEW', 'YEAR 2024', 'YEAR 2025', 'YEAR 2026', 'YEAR 2027', 'YEAR 2028', 'PLAYER DETAILS']:
                    for cell in row:
                        cell.font = Font(bold=True)
        
        print(f"\nSummary table saved to: {filename}")
    
    return df

def create_html_table(df, package_name):
    """
    Creates a formatted HTML table with simplified styling.
    
    Args:
        df (pandas.DataFrame): The data to format
        package_name (str): Name of the trade package
        
    Returns:
        str: Formatted HTML string
    """
    html = f"""
    <!DOCTYPE html>
    <html>
    <head>
        <title>Trade Package Summary - {package_name}</title>
        <style>
            body {{
                font-family: 'Segoe UI', Tahoma, Geneva, Verdana, sans-serif;
                margin: 20px;
                background-color: #f8f9fa;
            }}
            .container {{
                max-width: 1200px;
                margin: 0 auto;
                background-color: white;
                border-radius: 6px;
                box-shadow: 0 1px 3px rgba(0,0,0,0.1);
                overflow: hidden;
            }}
            .header {{
                background-color: #495057;
                color: white;
                padding: 20px;
                text-align: center;
            }}
            .header h1 {{
                margin: 0;
                font-size: 24px;
                font-weight: 400;
            }}
            .header p {{
                margin: 5px 0 0 0;
                opacity: 0.9;
                font-size: 14px;
            }}
            table {{
                width: 100%;
                border-collapse: collapse;
                font-size: 14px;
            }}
            th {{
                background-color: #6c757d;
                color: white;
                padding: 12px 8px;
                text-align: left;
                font-weight: 600;
                border: 1px solid #5a6268;
            }}
            td {{
                padding: 10px 8px;
                border: 1px solid #dee2e6;
                vertical-align: top;
            }}
            tr:nth-child(even) {{
                background-color: #f8f9fa;
            }}
            tr:hover {{
                background-color: #e9ecef;
            }}
            .category-header {{
                background-color: #495057;
                color: white;
                font-weight: bold;
                font-size: 16px;
                text-align: center;
            }}
            .category-header td {{
                padding: 15px 8px;
                border: 1px solid #343a40;
            }}
            .spacer-row td {{
                background-color: #f8f9fa;
                border: none;
                height: 10px;
            }}
            .metric-cell {{
                font-weight: 600;
                color: #212529;
                min-width: 200px;
            }}
            .value-cell {{
                color: #495057;
                min-width: 300px;
            }}
            .details-cell {{
                color: #6c757d;
                font-size: 13px;
                min-width: 250px;
            }}
            .positive {{
                color: #28a745;
                font-weight: 600;
            }}
            .negative {{
                color: #dc3545;
                font-weight: 600;
            }}
            .currency {{
                font-family: 'Courier New', monospace;
                font-weight: 600;
            }}
            .war-value {{
                font-weight: 600;
                color: #495057;
            }}
            .footer {{
                background-color: #f8f9fa;
                padding: 15px;
                text-align: center;
                color: #6c757d;
                font-size: 12px;
                border-top: 1px solid #dee2e6;
            }}
        </style>
    </head>
    <body>
        <div class="container">
            <div class="header">
                <h1>Trade Package Summary</h1>
                <p>{package_name}</p>
            </div>
            <table>
                <thead>
                    <tr>
                        <th>Category</th>
                        <th>Metric</th>
                        <th>Value</th>
                        <th>Details</th>
                    </tr>
                </thead>
                <tbody>
    """
    
    for _, row in df.iterrows():
        category = row['Category']
        metric = row['Metric']
        value = row['Value']
        details = row['Details']
        
        # Determine row styling
        if category == '':
            # Spacer row
            html += '<tr class="spacer-row"><td colspan="4"></td></tr>'
        elif category in ['PACKAGE OVERVIEW', 'YEAR 2024', 'YEAR 2025', 'YEAR 2026', 'YEAR 2027', 'YEAR 2028', 'PLAYER DETAILS']:
            # Category header
            html += f'<tr class="category-header"><td colspan="4">{category}</td></tr>'
        else:
            # Regular data row
            # Apply special formatting to values
            formatted_value = value
            if '$' in value:
                formatted_value = f'<span class="currency">{value}</span>'
            elif any(metric_word in metric.lower() for metric_word in ['war', 'roster', 'projected']):
                formatted_value = f'<span class="war-value">{value}</span>'
            
            # Apply color coding to net values
            if 'Net:' in details:
                if '+' in details:
                    details = f'<span class="positive">{details}</span>'
                elif '-' in details:
                    details = f'<span class="negative">{details}</span>'
            
            html += f'''
                <tr>
                    <td class="metric-cell">{category}</td>
                    <td class="metric-cell">{metric}</td>
                    <td class="value-cell">{formatted_value}</td>
                    <td class="details-cell">{details}</td>
                </tr>
            '''
    
    html += '''
                </tbody>
            </table>
            <div class="footer">
                <p>Generated by Trade Package Analysis Tool | Data from Notion API</p>
            </div>
        </div>
    </body>
    </html>
    '''
    
    return html

def generate_player_summary_html(package_data):
    """
    Generate a comprehensive player summary table for the trade package.
    
    Args:
        package_data (dict): Package data from pull_trade_package_data
        
    Returns:
        str: HTML string for the player summary table
    """
    if not package_data or not package_data["players"]:
        return ""
    
    # Calculate player statistics
    player_stats = []
    
    for player in package_data["players"]:
        player_name = player.get("name", "Unknown")
        organization = player.get("organization", "Unknown")
        position = player.get("position", "Unknown")
        
        # Calculate years of control (count unique years in valuations)
        years = set()
        total_war_proj = 0
        total_war_roster = 0
        total_salary = 0
        salary_2025 = 0
        asset_value = player.get("asset_value", 0)
        
        for val in player.get("valuations", []):
            year = val.get("year")
            if year:
                years.add(year)
                total_war_proj += val.get("war", 0)
                total_war_roster += val.get("war_roster", 0)
                total_salary += val.get("salary", 0)
                if year == 2025:
                    salary_2025 = val.get("salary", 0)
        
        years_of_control = len(years)
        
        # Do NOT divide by 1,000,000; values are already in millions
        salary_2025_millions = salary_2025
        
        player_stats.append({
            "name": player_name,
            "name_html": player.get("name_html", player_name),  # Include the HTML version with links
            "organization": organization,
            "position": position,
            "years_of_control": years_of_control,
            "total_war_proj": total_war_proj,
            "total_war_roster": total_war_roster,
            "asset_value": asset_value,
            "salary_2025": salary_2025_millions
        })
    
    # Sort by organization (PIT first) then by total WAR
    player_stats.sort(key=lambda x: (x["organization"] != "PIT", -x["total_war_roster"]))
    
    # Generate HTML table
    html = '''
        <div class="section">
            <div class="section-title">Player Summary</div>
            <table>
                <thead>
                    <tr>
                        <th>Player</th>
                        <th>Org</th>
                        <th>Pos</th>
                        <th>Years of Control</th>
                        <th>Total WAR Proj</th>
                        <th>Total WAR Adjusted</th>
                        <th>2025 Salary ($M)</th>
                        <th>Asset Value ($M)</th>
                    </tr>
                </thead>
                <tbody>
    '''
    
    # Group by organization
    pit_players = [p for p in player_stats if p["organization"] == "PIT"]
    non_pit_players = [p for p in player_stats if p["organization"] != "PIT"]
    
    # Add PIT players (Give)
    if pit_players:
        html += '<tr class="org-header"><td colspan="8" style="background-color: #FFD700; color: black; font-weight: bold; text-align: center;">GIVE</td></tr>'
        for player in pit_players:
            html += f'''
                <tr>
                    <td style="font-weight: 600; min-width: 120px; max-width: 150px; text-align: left; white-space: nowrap; overflow: hidden; text-overflow: ellipsis;">{player.get("name_html", player["name"])}</td>
                    <td>{player["organization"]}</td>
                    <td>{player["position"]}</td>
                    <td>{player["years_of_control"]}</td>
                    <td style="font-weight: 600;">{player["total_war_proj"]:.2f}</td>
                    <td style="font-weight: 600;">{player["total_war_roster"]:.2f}</td>
                    <td style="font-weight: 600; color: black;">${player["salary_2025"]:.1f}M</td>
                    <td style="font-weight: 600; color: black;">${player["asset_value"]:.1f}M</td>
                </tr>
            '''
    
    # Add non-PIT players (Receive)
    if non_pit_players:
        html += '<tr class="org-header"><td colspan="8" style="background-color: #000000; color: white; font-weight: bold; text-align: center;">RECEIVE</td></tr>'
        for player in non_pit_players:
            html += f'''
                <tr>
                    <td style="font-weight: 600; min-width: 120px; max-width: 150px; text-align: left; white-space: nowrap; overflow: hidden; text-overflow: ellipsis;">{player.get("name_html", player["name"])}</td>
                    <td>{player["organization"]}</td>
                    <td>{player["position"]}</td>
                    <td>{player["years_of_control"]}</td>
                    <td style="font-weight: 600;">{player["total_war_proj"]:.2f}</td>
                    <td style="font-weight: 600;">{player["total_war_roster"]:.2f}</td>
                    <td style="font-weight: 600; color: black;">${player["salary_2025"]:.1f}M</td>
                    <td style="font-weight: 600; color: black;">${player["asset_value"]:.1f}M</td>
                </tr>
            '''
    
    # Calculate PIT vs non-PIT totals
    pit_war_proj = sum(p["total_war_proj"] for p in pit_players)
    pit_war_roster = sum(p["total_war_roster"] for p in pit_players)
    pit_asset_value = sum(p["asset_value"] for p in pit_players)
    pit_salary_2025 = sum(p["salary_2025"] for p in pit_players)
    
    non_pit_war_proj = sum(p["total_war_proj"] for p in non_pit_players)
    non_pit_war_roster = sum(p["total_war_roster"] for p in non_pit_players)
    non_pit_asset_value = sum(p["asset_value"] for p in non_pit_players)
    non_pit_salary_2025 = sum(p["salary_2025"] for p in non_pit_players)
    
    # Calculate net values for coloring
    net_asset_value = non_pit_asset_value - pit_asset_value  # Receive - Give
    net_salary_2025 = non_pit_salary_2025 - pit_salary_2025  # Non-PIT - PIT (positive = savings)
    net_war_proj = non_pit_war_proj - pit_war_proj  # Receive - Give
    net_war_roster = non_pit_war_roster - pit_war_roster  # Receive - Give
    
    # Determine colors based on the new logic
    asset_value_color = "#4CAF50" if net_asset_value >= 0 else "#F44336"  # Green if positive, red if negative
    salary_2025_color = "#4CAF50" if net_salary_2025 <= 0 else "#F44336"  # Green if negative (savings), red if positive (costs)
    war_proj_color = "#4CAF50" if net_war_proj >= 0 else "#F44336"  # Green if positive (gain), red if negative (loss)
    war_roster_color = "#4CAF50" if net_war_roster >= 0 else "#F44336"  # Green if positive (gain), red if negative (loss)
    
    html += f'''
        <tr class="net-row" style="border-top: 3px solid #343a40;">
            <td colspan="3" style="font-weight: bold;">NET (Receive - Give)</td>
            <td style="font-weight: bold;"></td>
            <td style="font-weight: bold; color: {war_proj_color};">{non_pit_war_proj - pit_war_proj:+.2f}</td>
            <td style="font-weight: bold; color: {war_roster_color};">{non_pit_war_roster - pit_war_roster:+.2f}</td>
            <td style="font-weight: bold; color: {salary_2025_color};">${net_salary_2025:+.1f}M</td>
            <td style="font-weight: bold; color: {asset_value_color};">${net_asset_value:+.1f}M</td>
        </tr>
    '''
    
    html += '''
                </tbody>
            </table>
        </div>
    '''
    # Add package notes if present
    notes = package_data.get('package', {}).get('notes') or package_data.get('package', {}).get('Notes') or package_data.get('package', {}).get('description') or package_data.get('package', {}).get('Description')
    if notes:
        html += f'''
        <div class="section" style="margin-top: 0;">
            <div class="section-title" style="font-size:16px; background:#e9ecef; color:#333;">Package Notes</div>
            <div style="padding: 12px 18px; background: #f8f9fa; border-radius: 4px; border: 1px solid #dee2e6; font-size: 14px; color: #333;">{notes}</div>
        </div>
        '''
    else:
        # Debug: Print available package fields
        print("DEBUG: No notes found. Available package fields:")
        for key in package_data.get('package', {}).keys():
            print(f"  - {key}: {package_data['package'][key]}")
        print("DEBUG: Package data structure:")
        print(package_data.get('package', {}))
    
    return html

def create_metric_comparison_html(df, package_name, include_plots=True, package_data=None, plot_type="all"):
    """
    Creates a formatted HTML table for metric comparison with optional plots.
    
    Args:
        df (pandas.DataFrame): The comparison data to format
        package_name (str): Name of the trade package
        include_plots (bool): Whether to include interactive plots
        package_data (dict): Package data for generating plots
        plot_type (str): Type of plots to show - "all" for all 4 plots, "war_adjusted" for just WAR Adjusted plot
        
    Returns:
        str: Formatted HTML string
    """
    # Generate plot data if requested
    plot_html = ""
    if include_plots and package_data:
        plot_html = generate_interactive_plots_html(package_data, plot_type)
    
    # Generate player summary table if package data is provided
    player_summary_html = ""
    if package_data:
        player_summary_html = generate_player_summary_html(package_data)
    
    html = f"""
    <!DOCTYPE html>
    <html>
    <head>
        <title>Trade Package Analysis - {package_name}</title>
        <script src="https://cdn.plot.ly/plotly-latest.min.js"></script>
        <style>
            body {{
                font-family: 'Segoe UI', Tahoma, Geneva, Verdana, sans-serif;
                margin: 20px;
                background-color: #f8f9fa;
            }}
            .container {{
                max-width: 1600px;
                margin: 0 auto;
                background-color: white;
                border-radius: 6px;
                box-shadow: 0 1px 3px rgba(0,0,0,0.1);
                overflow: hidden;
            }}
            .header {{
                background-color: #495057;
                color: white;
                padding: 20px;
                text-align: center;
            }}
            .header h1 {{
                margin: 0;
                font-size: 28px;
                font-weight: 400;
            }}
            .header p {{
                margin: 5px 0 0 0;
                opacity: 0.9;
                font-size: 16px;
            }}
            .section {{
                margin: 30px 0;
                padding: 0 20px;
            }}
            .section-title {{
                background-color: #6c757d;
                color: white;
                padding: 15px 20px;
                margin: 0 -20px 20px -20px;
                font-size: 20px;
                font-weight: 600;
            }}
            table {{
                width: 100%;
                border-collapse: collapse;
                font-size: 12px;
                margin-bottom: 30px;
            }}
            th {{
                background-color: #6c757d;
                color: white;
                padding: 10px 6px;
                text-align: center;
                font-weight: 600;
                border: 1px solid #5a6268;
                font-size: 11px;
            }}
            td {{
                padding: 8px 6px;
                border: 1px solid #dee2e6;
                text-align: center;
                vertical-align: middle;
            }}
            tr:nth-child(even) {{
                background-color: #f8f9fa;
            }}
            tr:hover {{
                background-color: #e9ecef;
            }}
            .pit-section {{
                background-color: #f8f9fa;
                font-weight: 600;
            }}
            .non-pit-section {{
                background-color: #e9ecef;
                font-weight: 600;
            }}
            .net-section {{
                background-color: #dee2e6;
                font-weight: bold;
            }}
            .total-row {{
                background-color: #343a40;
                color: white;
                font-weight: bold;
                font-size: 15px;
            }}
            .total-row td {{
                border: 1px solid #495057;
                padding: 12px 6px;
            }}
            .total-row .positive {{
                color: #90EE90;
                font-weight: bold;
            }}
            .total-row .negative {{
                color: #FFB6C1;
                font-weight: bold;
            }}
            .total-row .currency {{
                color: #FFD700;
                font-weight: bold;
            }}
            .total-row .war-value {{
                color: #87CEEB;
                font-weight: bold;
            }}
            .positive {{
                color: #28a745;
                font-weight: 600;
            }}
            .negative {{
                color: #dc3545;
                font-weight: 600;
            }}
            .currency {{
                font-family: 'Courier New', monospace;
                font-weight: 600;
            }}
            .war-value {{
                font-weight: 600;
                color: #495057;
            }}
            .plots-container {{
                display: grid;
                grid-template-columns: 1fr 1fr;
                gap: 20px;
                margin: 20px 0;
            }}
            .plot {{
                background-color: white;
                border: 1px solid #dee2e6;
                border-radius: 6px;
                padding: 15px;
                box-shadow: 0 1px 3px rgba(0,0,0,0.1);
            }}
            .org-header {{
                background-color: #f8f9fa;
                font-weight: bold;
                text-align: center;
                border: 1px solid #dee2e6;
            }}
            .net-row {{
                background-color: #e8f5e8;
                font-weight: bold;
                border: 1px solid #4CAF50;
            }}
            .net-row td {{
                padding: 10px 6px;
                border: 1px solid #4CAF50;
            }}
            .footer {{
                background-color: #f8f9fa;
                padding: 15px;
                text-align: center;
                color: #6c757d;
                font-size: 12px;
                border-top: 1px solid #dee2e6;
            }}
        </style>
    </head>
    <body>
        <div class="container">
            <div class="header">
                <h1>Trade Package Analysis</h1>
                <p>{package_name}</p>
            </div>
            
            {player_summary_html}
            
            <div class="section">
                <div class="section-title collapsible" onclick="toggleSection('metric-table')">
                    Metric Comparison Table <span class="toggle-icon">▼</span>
                </div>
                <div id="metric-table" class="collapsible-content">
                    <table>
                    <thead>
                        <tr>
                            <th rowspan="2">Year</th>
                            <th colspan="2">Players</th>
                            <th colspan="3">WAR Projected</th>
                            <th colspan="3">WAR Adjusted</th>
                            <th colspan="3">Expected Salary ($M)<br><small>Net = Non-PIT - PIT</small></th>
                            <th colspan="3">Asset Value ($M)</th>
                        </tr>
                        <tr>
                            <th>Give</th>
                            <th>Receive</th>
                            <th>Give</th>
                            <th>Receive</th>
                            <th>Net</th>
                            <th>Give</th>
                            <th>Receive</th>
                            <th>Net</th>
                            <th>Give</th>
                            <th>Receive</th>
                            <th>Net</th>
                            <th>Give</th>
                            <th>Receive</th>
                            <th>Net</th>
                        </tr>
                    </thead>
                    <tbody>
    """
    
    for _, row in df.iterrows():
        year = row['Year']
        
        # Determine if this is the total row
        is_total_row = year == 'TOTAL'
        row_style = ' style="border-top: 3px solid #343a40;"' if is_total_row else ''
        
        # Format values with color coding
        pit_war_proj = f'<span class="war-value">{row["Give WAR Proj"]}</span>'
        non_pit_war_proj = f'<span class="war-value">{row["Receive WAR Proj"]}</span>'
        net_war_proj = row["Net WAR Proj"]
        if '+' in net_war_proj:
            net_war_proj = f'<span class="positive">{net_war_proj}</span>'
        elif '-' in net_war_proj:
            net_war_proj = f'<span class="negative">{net_war_proj}</span>'
        else:
            net_war_proj = f'<span class="war-value">{net_war_proj}</span>'
        
        pit_war_roster = f'<span class="war-value">{row["Give WAR Adjusted"]}</span>'
        non_pit_war_roster = f'<span class="war-value">{row["Receive WAR Adjusted"]}</span>'
        net_war_roster = row["Net WAR Adjusted"]
        if '+' in net_war_roster:
            net_war_roster = f'<span class="positive">{net_war_roster}</span>'
        elif '-' in net_war_roster:
            net_war_roster = f'<span class="negative">{net_war_roster}</span>'
        else:
            net_war_roster = f'<span class="war-value">{net_war_roster}</span>'
        
        pit_salary = f'<span class="currency">{row["Give Salary"]}</span>'
        non_pit_salary = f'<span class="currency">{row["Receive Salary"]}</span>'
        net_salary = row["Net Salary"]
        if '+' in net_salary:
            net_salary = f'<span class="negative">{net_salary}</span>'  # Red for positive (costs)
        elif '-' in net_salary:
            net_salary = f'<span class="positive">{net_salary}</span>'  # Green for negative (savings)
        else:
            net_salary = f'<span class="currency">{net_salary}</span>'
        
        pit_av = f'<span class="currency">{row["Give AV"]}</span>'
        non_pit_av = f'<span class="currency">{row["Receive AV"]}</span>'
        net_av = row["Net AV"]
        if '+' in net_av:
            net_av = f'<span class="positive">{net_av}</span>'  # Green for positive (gain)
        elif '-' in net_av:
            net_av = f'<span class="negative">{net_av}</span>'  # Red for negative (loss)
        else:
            net_av = f'<span class="currency">{net_av}</span>'
        
        html += f'''
            <tr{row_style}>
                <td class="pit-section">{year}</td>
                <td class="pit-section">{row["Give Players"]}</td>
                <td class="non-pit-section">{row["Receive Players"]}</td>
                <td class="pit-section">{pit_war_proj}</td>
                <td class="non-pit-section">{non_pit_war_proj}</td>
                <td class="net-section">{net_war_proj}</td>
                <td class="pit-section">{pit_war_roster}</td>
                <td class="non-pit-section">{non_pit_war_roster}</td>
                <td class="net-section">{net_war_roster}</td>
                <td class="pit-section">{pit_salary}</td>
                <td class="non-pit-section">{non_pit_salary}</td>
                <td class="net-section">{net_salary}</td>
                <td class="pit-section">{pit_av}</td>
                <td class="non-pit-section">{non_pit_av}</td>
                <td class="net-section">{net_av}</td>
            </tr>
        '''
    
    html += '''
                </tbody>
            </table>
                </div>
        </div>
    '''
    # Add plots section if plots are included
    if plot_html:
        html += f'''
        <div class="section">
            <div class="section-title">Interactive Analysis Plots</div>
            {plot_html}
        </div>
        '''
    html += '''
            <div class="footer">
                <p>Generated by Trade Package Analysis Tool | Data from Notion API</p>
            </div>
        </div>
    </body>
    <script>
    function toggleSection(id) {
        var content = document.getElementById(id);
        var icon = content.previousElementSibling.querySelector('.toggle-icon');
        if (content.style.display === 'none') {
            content.style.display = 'block';
            icon.textContent = '▼';
        } else {
            content.style.display = 'none';
            icon.textContent = '►';
        }
    }
    // Start collapsed
    document.addEventListener('DOMContentLoaded', function() {
        var content = document.getElementById('metric-table');
        if (content) {
            content.style.display = 'none';
            var icon = content.previousElementSibling.querySelector('.toggle-icon');
            if (icon) icon.textContent = '►';
        }
    });
    </script>
    <style>
    .collapsible { cursor: pointer; user-select: none; }
    .toggle-icon { font-size: 1em; margin-left: 8px; }
    .collapsible-content { transition: all 0.2s; }
    </style>
    </html>
    '''
    return html

def create_metric_comparison_table(package_data, save_to_excel=False, filename=None, output_html=True):
    """
    Creates a focused metric comparison table showing key differences.
    
    Args:
        package_data (dict): Output from pull_trade_package_data function
        save_to_excel (bool): Whether to save the table to an Excel file
        filename (str): Optional filename for Excel export
        output_html (bool): Whether to output as formatted HTML
        
    Returns:
        pandas.DataFrame: Formatted comparison table
    """
    if not package_data or not package_data["players"]:
        print("No package data available for table creation")
        return None
    
    # Group valuations by year and organization
    year_org_data = {}
    
    for player in package_data["players"]:
        org = player.get("organization", "Unknown")
        print(f"DEBUG: Player {player.get('name', 'Unknown')} has org: {org}")
        for val in player.get("valuations", []):
            year = val.get("year")
            salary = val.get("salary", 0)
            print(f"DEBUG: Year {year}, salary: {salary}")
            if year:
                if year not in year_org_data:
                    year_org_data[year] = {
                        "PIT": {"war_roster": 0, "war_proj": 0, "salary": 0, "av": 0, "players": 0}, 
                        "non_PIT": {"war_roster": 0, "war_proj": 0, "salary": 0, "av": 0, "players": 0}
                    }
                
                if org == "PIT":
                    year_org_data[year]["PIT"]["war_roster"] += val.get("war_roster", 0)
                    year_org_data[year]["PIT"]["war_proj"] += val.get("war", 0)
                    year_org_data[year]["PIT"]["salary"] += val.get("salary", 0)
                    year_org_data[year]["PIT"]["av"] += val.get("asset_value", 0)
                    year_org_data[year]["PIT"]["players"] += 1
                    print(f"DEBUG: Added to PIT - Year {year} salary now: {year_org_data[year]['PIT']['salary']}")
                else:
                    year_org_data[year]["non_PIT"]["war_roster"] += val.get("war_roster", 0)
                    year_org_data[year]["non_PIT"]["war_proj"] += val.get("war", 0)
                    year_org_data[year]["non_PIT"]["salary"] += val.get("salary", 0)
                    year_org_data[year]["non_PIT"]["av"] += val.get("asset_value", 0)
                    year_org_data[year]["non_PIT"]["players"] += 1
                    print(f"DEBUG: Added to non-PIT - Year {year} salary now: {year_org_data[year]['non_PIT']['salary']}")
    
    # Create comparison table
    comparison_data = []
    
    # Calculate totals across all years
    total_pit_players = 0
    total_non_pit_players = 0
    total_pit_war_proj = 0
    total_non_pit_war_proj = 0
    total_pit_war_adjusted = 0
    total_non_pit_war_adjusted = 0
    total_pit_salary = 0
    total_non_pit_salary = 0
    total_pit_av = 0
    total_non_pit_av = 0
    
    for year in sorted(year_org_data.keys()):
        pit_data = year_org_data[year]["PIT"]
        non_pit_data = year_org_data[year]["non_PIT"]
        
        # Net calculations
        net_war_roster = non_pit_data['war_roster'] - pit_data['war_roster']  # Non-PIT - PIT (positive = gain for non-PIT)
        net_war_proj = non_pit_data['war_proj'] - pit_data['war_proj']  # Non-PIT - PIT (positive = gain for non-PIT)
        # Do NOT divide by 1,000,000; values are already in millions
        pit_salary_millions = pit_data['salary']
        non_pit_salary_millions = non_pit_data['salary']
        net_salary = non_pit_salary_millions - pit_salary_millions  # Non-PIT - PIT (positive = savings)
        pit_av_millions = pit_data['av']
        non_pit_av_millions = non_pit_data['av']
        net_av = non_pit_av_millions - pit_av_millions  # Non-PIT - PIT (positive = gain for non-PIT)
        
        comparison_data.append({
            'Year': year,
            'Give Players': pit_data['players'],
            'Receive Players': non_pit_data['players'],
            'Give WAR Proj': f"{pit_data['war_proj']:.2f}",
            'Receive WAR Proj': f"{non_pit_data['war_proj']:.2f}",
            'Net WAR Proj': f"{net_war_proj:+.2f}",
            'Give WAR Adjusted': f"{pit_data['war_roster']:.2f}",
            'Receive WAR Adjusted': f"{non_pit_data['war_roster']:.2f}",
            'Net WAR Adjusted': f"{net_war_roster:+.2f}",
            'Give Salary': f"${pit_salary_millions:.1f}M",
            'Receive Salary': f"${non_pit_salary_millions:.1f}M",
            'Net Salary': f"${net_salary:+.1f}M",
            'Give AV': f"${pit_av_millions:.1f}M",
            'Receive AV': f"${non_pit_av_millions:.1f}M",
            'Net AV': f"${net_av:+.1f}M"
        })
        
        # Add to totals
        total_pit_players += pit_data['players']
        total_non_pit_players += non_pit_data['players']
        total_pit_war_proj += pit_data['war_proj']
        total_non_pit_war_proj += non_pit_data['war_proj']
        total_pit_war_adjusted += pit_data['war_roster']
        total_non_pit_war_adjusted += non_pit_data['war_roster']
        total_pit_salary += pit_data['salary']
        total_non_pit_salary += non_pit_data['salary']
        total_pit_av += pit_data['av']
        total_non_pit_av += non_pit_data['av']
    
    # Calculate net totals
    total_net_war_proj = total_non_pit_war_proj - total_pit_war_proj  # Non-PIT - PIT (positive = gain for non-PIT)
    total_net_war_adjusted = total_non_pit_war_adjusted - total_pit_war_adjusted  # Non-PIT - PIT (positive = gain for non-PIT)
    # Do NOT divide by 1,000,000; values are already in millions
    total_pit_salary_millions = total_pit_salary
    total_non_pit_salary_millions = total_non_pit_salary
    total_net_salary = total_non_pit_salary_millions - total_pit_salary_millions  # Non-PIT - PIT (positive = savings)
    total_pit_av_millions = total_pit_av
    total_non_pit_av_millions = total_non_pit_av
    total_net_av = total_non_pit_av_millions - total_pit_av_millions  # Non-PIT - PIT (positive = gain for non-PIT)
    
    # Add total row
    comparison_data.append({
        'Year': 'TOTAL',
        'Give Players': total_pit_players,
        'Receive Players': total_non_pit_players,
        'Give WAR Proj': f"{total_pit_war_proj:.2f}",
        'Receive WAR Proj': f"{total_non_pit_war_proj:.2f}",
        'Net WAR Proj': f"{total_net_war_proj:+.2f}",
        'Give WAR Adjusted': f"{total_pit_war_adjusted:.2f}",
        'Receive WAR Adjusted': f"{total_non_pit_war_adjusted:.2f}",
        'Net WAR Adjusted': f"{total_net_war_adjusted:+.2f}",
        'Give Salary': f"${total_pit_salary_millions:.1f}M",
        'Receive Salary': f"${total_non_pit_salary_millions:.1f}M",
        'Net Salary': f"${total_net_salary:+.1f}M",
        'Give AV': f"${total_pit_av_millions:.1f}M",
        'Receive AV': f"${total_non_pit_av_millions:.1f}M",
        'Net AV': f"${total_net_av:+.1f}M"
    })
    
    # Create DataFrame
    df = pd.DataFrame(comparison_data)
    
    # Reorder columns to move Asset Value to the end
    column_order = [
        'Year', 'Give Players', 'Receive Players',
        'Give WAR Proj', 'Receive WAR Proj', 'Net WAR Proj',
        'Give WAR Adjusted', 'Receive WAR Adjusted', 'Net WAR Adjusted',
        'Give Salary', 'Receive Salary', 'Net Salary',
        'Give AV', 'Receive AV', 'Net AV'
    ]
    df = df[column_order]
    
    # Output HTML table if requested
    if output_html:
        html_output = create_metric_comparison_html(df, package_data["package"].get("name", "Unknown"))
        print(html_output)
        
        # Save HTML to file
        html_filename = f"metric_comparison_{package_data['package'].get('name', 'unknown').replace(' ', '_')}.html"
        with open(html_filename, 'w', encoding='utf-8') as f:
            f.write(html_output)
        print(f"\nHTML comparison table saved to: {html_filename}")
    
    # Display the table
    print("=" * 150)
    print(f"METRIC COMPARISON TABLE")
    print(f"Package: {package_data['package'].get('name', 'Unknown')}")
    print("=" * 150)
    
    print(df.to_string(index=False))
    print("\n" + "=" * 150)
    
    # Save to Excel if requested
    if save_to_excel:
        if filename is None:
            filename = f"metric_comparison_{package_data['package'].get('name', 'unknown').replace(' ', '_')}.xlsx"
        
        with pd.ExcelWriter(filename, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='Comparison', index=False)
            
            # Auto-adjust column widths
            workbook = writer.book
            worksheet = writer.sheets['Comparison']
            
            for column in worksheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 20)
                worksheet.column_dimensions[column_letter].width = adjusted_width
            
            # Simple header formatting - just bold
            from openpyxl.styles import Font
            for cell in worksheet[1]:
                cell.font = Font(bold=True)
        
        print(f"\nComparison table saved to: {filename}")
    
    return df

def generate_interactive_plots_html(package_data, plot_type="all"):
    """
    Generate Plotly HTML divs for WAR Projected, WAR Adjusted, Salary, and Asset Value.
    Returns a string of HTML divs in a 2x2 grid or single plot based on plot_type.
    
    Args:
        package_data (dict): Package data for generating plots
        plot_type (str): Type of plots to show - "all" for all 4 plots, "war_adjusted" for just WAR Adjusted plot
    """
    # Create detailed player breakdown data for tooltips
    year_player_data = {}
    for player in package_data["players"]:
        player_name = player.get("name", "Unknown")
        org = player.get("organization", "Unknown")
        position = player.get("position", "Unknown")
        
        for val in player.get("valuations", []):
            year = val.get("year")
            if year:
                if year not in year_player_data:
                    year_player_data[year] = {"PIT": [], "non_PIT": []}
                
                player_data = {
                    "name": player_name,
                    "position": position,
                    "war_roster": val.get("war_roster", 0),
                    "war_proj": val.get("war", 0),
                    "salary": val.get("salary", 0),
                    "av": val.get("asset_value", 0)
                }
                
                if org == "PIT":
                    year_player_data[year]["PIT"].append(player_data)
                else:
                    year_player_data[year]["non_PIT"].append(player_data)
    
    # Group valuations by year and organization for totals
    year_org_data = {}
    for player in package_data["players"]:
        org = player.get("organization", "Unknown")
        for val in player.get("valuations", []):
            year = val.get("year")
            if year:
                if year not in year_org_data:
                    year_org_data[year] = {
                        "PIT": {"war_roster": 0, "war_proj": 0, "salary": 0, "av": 0},
                        "non_PIT": {"war_roster": 0, "war_proj": 0, "salary": 0, "av": 0}
                    }
                if org == "PIT":
                    year_org_data[year]["PIT"]["war_roster"] += val.get("war_roster", 0)
                    year_org_data[year]["PIT"]["war_proj"] += val.get("war", 0)
                    year_org_data[year]["PIT"]["salary"] += val.get("salary", 0)
                    year_org_data[year]["PIT"]["av"] += val.get("asset_value", 0)
                else:
                    year_org_data[year]["non_PIT"]["war_roster"] += val.get("war_roster", 0)
                    year_org_data[year]["non_PIT"]["war_proj"] += val.get("war", 0)
                    year_org_data[year]["non_PIT"]["salary"] += val.get("salary", 0)
                    year_org_data[year]["non_PIT"]["av"] += val.get("asset_value", 0)
    
    years = sorted(year_org_data.keys())
    
    def create_tooltip_text(year, org, metric):
        """Create detailed tooltip text showing player breakdowns"""
        if year not in year_player_data:
            return f"<b>{org}</b><br>Total: {year_org_data[year][org][metric]:.2f}"
        
        players = year_player_data[year][org]
        if not players:
            return f"<b>{org}</b><br>Total: {year_org_data[year][org][metric]:.2f}"
        
        # Sort players by the metric value (descending)
        players_sorted = sorted(players, key=lambda x: x[metric], reverse=True)
        
        tooltip = f"<b>{org}</b><br>"
        
        # Add total with appropriate units
        if metric == "salary":
            tooltip += f"<b>Total: ${year_org_data[year][org][metric]:.1f}M</b><br><br>"
        elif metric == "av":
            tooltip += f"<b>Total: ${year_org_data[year][org][metric]:.1f}M</b><br><br>"
        elif metric == "war_roster":
            tooltip += f"<b>Total: {year_org_data[year][org][metric]:.2f} WAR</b><br><br>"
        elif metric == "war_proj":
            tooltip += f"<b>Total: {year_org_data[year][org][metric]:.2f} WAR</b><br><br>"
        else:
            tooltip += f"<b>Total: {year_org_data[year][org][metric]:.2f}</b><br><br>"
        
        for player in players_sorted:
            if metric == "salary":
                tooltip += f"• {player['name']} ({player['position']}): ${player[metric]:.1f}M<br>"
            elif metric == "av":
                tooltip += f"• {player['name']} ({player['position']}): ${player[metric]:.1f}M<br>"
            elif metric == "war_roster":
                tooltip += f"• {player['name']} ({player['position']}): {player[metric]:.2f} WAR (${player['salary']:.1f}M)<br>"
            elif metric == "war_proj":
                tooltip += f"• {player['name']} ({player['position']}): {player[metric]:.2f} WAR (${player['salary']:.1f}M)<br>"
            else:
                tooltip += f"• {player['name']} ({player['position']}): {player[metric]:.2f}<br>"
        
        return tooltip
    
    def get_lists_with_tooltips(metric):
        pit_values = []
        non_pit_values = []
        pit_tooltips = []
        non_pit_tooltips = []
        
        for year in years:
            pit_values.append(year_org_data[year]["PIT"][metric])
            non_pit_values.append(year_org_data[year]["non_PIT"][metric])
            pit_tooltips.append(create_tooltip_text(year, "PIT", metric))
            non_pit_tooltips.append(create_tooltip_text(year, "non_PIT", metric))
        
        return pit_values, non_pit_values, pit_tooltips, non_pit_tooltips
    
    plots = []
    
    if plot_type == "war_adjusted":
        # Only generate WAR Adjusted plot with detailed tooltips
        metric, label, yformat = ("war_roster", "WAR Adjusted", None)
        pit, non_pit, pit_tooltips, non_pit_tooltips = get_lists_with_tooltips(metric)
        
        fig = go.Figure()
        
        # Add lines only (no markers) for all points
        fig.add_trace(go.Scatter(
            x=years, y=pit, mode='lines', name='Give', 
            line=dict(color='blue', width=3), 
            showlegend=True,
            hovertemplate='%{text}<extra></extra>',
            text=pit_tooltips
        ))
        fig.add_trace(go.Scatter(
            x=years, y=non_pit, mode='lines', name='Receive', 
            line=dict(color='red', width=3), 
            showlegend=True,
            hovertemplate='%{text}<extra></extra>',
            text=non_pit_tooltips
        ))
        
        fig.update_layout(
            title=label, 
            xaxis_title='Year', 
            yaxis_title=label, 
            height=400, 
            margin=dict(l=40, r=20, t=40, b=40), 
            legend=dict(orientation="h", yanchor="bottom", y=1.02, xanchor="right", x=1),
            hovermode='closest'
        )
        
        if yformat:
            fig.update_yaxes(tickformat=yformat)
        
        plots.append(plot(fig, output_type='div', include_plotlyjs=False))
        # Single plot layout
        html = '<div class="plots-container" style="grid-template-columns: 1fr;">' + ''.join(f'<div class="plot">{div}</div>' for div in plots) + '</div>'
    else:
        # Generate all 4 plots with detailed tooltips
        for metric, label, yformat in [
            ("war_proj", "WAR Projected", None),
            ("war_roster", "WAR Adjusted", None),
            ("salary", "Expected Salary ($M)", ",.1f"),
            ("av", "Asset Value ($M)", ",.1f")
        ]:
            pit, non_pit, pit_tooltips, non_pit_tooltips = get_lists_with_tooltips(metric)
            
            fig = go.Figure()
            fig.add_trace(go.Scatter(
                x=years, y=pit, mode='lines+markers', name='Give', 
                line=dict(color='blue'), marker=dict(size=6),
                hovertemplate='%{text}<extra></extra>',
                text=pit_tooltips
            ))
            fig.add_trace(go.Scatter(
                x=years, y=non_pit, mode='lines+markers', name='Receive', 
                line=dict(color='red'), marker=dict(size=6),
                hovertemplate='%{text}<extra></extra>',
                text=non_pit_tooltips
            ))
            
            fig.update_layout(
                title=label, 
                xaxis_title='Year', 
                yaxis_title=label, 
                height=350, 
                margin=dict(l=40, r=20, t=40, b=40), 
                legend=dict(orientation="h", yanchor="bottom", y=1.02, xanchor="right", x=1),
                hovermode='closest'
            )
            
            if yformat:
                fig.update_yaxes(tickformat=yformat)
            
            plots.append(plot(fig, output_type='div', include_plotlyjs=False))
        
        # 2x2 grid
        html = '<div class="plots-container">' + ''.join(f'<div class="plot">{div}</div>' for div in plots) + '</div>'
    
    # Add plotly.js script ONCE at the top
    html = '<script src="https://cdn.plot.ly/plotly-latest.min.js"></script>' + html
    return html

if package_id is None:
    pkg = example_get_package_by_name(package_name)
else:
    pkg =  pull_trade_package_data(package_id)

# pkg_analysis = analyze_trade_package(pkg)
# print_trade_package_summary(pkg, pkg_analysis)


# print_war_comparison_summary(compare_war_projections_vs_roster(pkg))

# plot_metric_changes_over_time(pkg, metric='war_roster')
# plot_all_metrics_comparison(pkg)



df = create_metric_comparison_table(pkg, save_to_excel=False, filename='metric_comparison_Bednar_to_TB.xlsx', output_html=False)
html = create_metric_comparison_html(df, package_name= package_name, include_plots=True, package_data=pkg, plot_type="all")

# Save the combined HTML to a file
#with open("D:\$awan\KT\Reports\\" + package_name + " Trade Summary.html", 'w', encoding='utf-8') as f:
#    f.write(html)
print("Combined HTML file saved:" +  package_name + " Trade Summary.html")


# Notes
# Remove total salary column
# Add a final column at the end showing guaranteed money change (if there is any). Don't do this if it's too complicated